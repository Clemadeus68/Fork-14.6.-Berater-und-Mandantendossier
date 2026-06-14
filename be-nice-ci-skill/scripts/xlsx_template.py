"""
be nice network — Excel-Vorlage (.xlsx)
CI-konforme Tabellenvorlage für Clemens Gutmann
Nutzt openpyxl zur Formatierung nach be-nice-Standard.

VERWENDUNG:
  1. METADATEN unten anpassen (Titel, Datum)
  2. DATEN unten anpassen (sheets Dict mit Blättern und Zeilen)
  3. Ausführen: python3 xlsx_template.py

VORAUSSETZUNGEN:
  pip install openpyxl pillow --break-system-packages
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

# ── Farben ─────────────────────────────────────────────────────────────────────
GRUEN        = "8CC63E"
DUNKELGRAU   = "454544"
TUERKIS      = "33AB97"
BLAU_TUERKIS = "13A1B6"
HELLGRAU     = "E9E9E9"
WEISS        = "FFFFFF"
FAST_SCHWARZ = "111111"
MITTELGRAU   = "777777"

# ── Schriftarten ───────────────────────────────────────────────────────────────
FONT_NAME = "Calibri"
FONT_SIZE_TITLE = 18
FONT_SIZE_HEADER = 11
FONT_SIZE_DATA = 10
FONT_SIZE_META = 9

# ── Metadaten ──────────────────────────────────────────────────────────────────
META = {
    "titel":      "Excel Tabelle",
    "untertitel": "be nice Corporate Design",
    "datum":      "Juni 2026",
    "autor":      "Clemens Gutmann · be nice network",
    "ausgabe":    "Tabelle_Titel.xlsx",
}

# ── Styles ─────────────────────────────────────────────────────────────────────

def style_title(cell):
    """Titel-Zelle: dunkelgrau, fett, große Schrift."""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE_TITLE, bold=True, color=DUNKELGRAU)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_header(cell):
    """Header-Zelle: grüner Hintergrund, weiße Schrift, fett."""
    cell.fill = PatternFill(start_color=GRUEN, end_color=GRUEN, fill_type="solid")
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE_HEADER, bold=True, color=WEISS)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=MITTELGRAU),
        right=Side(style="thin", color=MITTELGRAU),
        top=Side(style="thin", color=MITTELGRAU),
        bottom=Side(style="thin", color=MITTELGRAU)
    )

def style_data(cell, is_total=False):
    """Datenzelle: hellgrauer Hintergrund, dunkle Schrift."""
    if is_total:
        cell.fill = PatternFill(start_color=TUERKIS, end_color=TUERKIS, fill_type="solid")
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE_HEADER, bold=True, color=WEISS)
    else:
        cell.fill = PatternFill(start_color=HELLGRAU, end_color=HELLGRAU, fill_type="solid")
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE_DATA, color=FAST_SCHWARZ)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

def style_meta(cell):
    """Metazeile: grau, kleine Schrift."""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE_META, color=MITTELGRAU, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

# ══════════════════════════════════════════════════════════════════════════════
# DATEN — hier anpassen
# ══════════════════════════════════════════════════════════════════════════════

sheets = {
    "Daten": [
        # Header-Zeile (wird grün gefärbt)
        ["Kategorie", "Q1", "Q2", "Q3", "Q4"],
        # Datenzeilen (werden grau gefärbt)
        ["Beratung", "12.500", "14.200", "13.800", "15.600"],
        ["Schulung", "8.300", "9.100", "8.950", "10.200"],
        ["Support", "5.600", "5.900", "6.200", "6.800"],
        # Summe-Zeile (wird türkis gefärbt)
        ["Gesamt", "26.400", "29.200", "28.950", "32.600"],
    ]
}

# ══════════════════════════════════════════════════════════════════════════════
# ARBEITSMAPPE ERSTELLEN
# ══════════════════════════════════════════════════════════════════════════════

wb = Workbook()
ws = wb.active
ws.title = "Daten"

# Titel-Zeile (fusioniert über alle Spalten)
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = META["titel"]
style_title(title_cell)
ws.row_dimensions[1].height = 25

# Meta-Zeile (fusioniert)
ws.merge_cells("A2:E2")
meta_cell = ws["A2"]
meta_cell.value = f'{META["autor"]} · {META["datum"]}'
style_meta(meta_cell)
ws.row_dimensions[2].height = 15

# Leere Zeile
ws.row_dimensions[3].height = 8

# Daten mit Styles
data = sheets["Daten"]
row_offset = 4  # Datenbeginn nach Titel und Meta

for row_idx, row_data in enumerate(data, start=row_offset):
    is_header = (row_idx == row_offset)
    is_total = (row_data[0] == "Gesamt")

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value

        if is_header:
            style_header(cell)
        elif is_total:
            style_data(cell, is_total=True)
        else:
            style_data(cell)

    ws.row_dimensions[row_idx].height = 20

# Spaltenbreiten
ws.column_dimensions["A"].width = 16
for col in ["B", "C", "D", "E"]:
    ws.column_dimensions[col].width = 14

# Arbeitsmappe speichern
wb.save(META["ausgabe"])
print(f"✓ Gespeichert: {META['ausgabe']}")
